import openpyxl as op


def get_row_count(file, sheetname):
    wb = op.load_workbook(file)
    sheet = wb[sheetname]
    return sheet.max_row


def get_col_count(file, sheetname):
    wb = op.load_workbook(file)
    sheet = wb[sheetname]
    return sheet.max_column


def read_data_by_row_col(file, sheetname, row_number, col_number):
    wb = op.load_workbook(file)
    sheet = wb[sheetname]
    return sheet.cell(row=row_number, column=col_number).value


def read_data_by_row(file, sheetname, row_number):
    wb = op.load_workbook(file)
    sheet = wb[sheetname]
    data_key = list()
    #  get value in row number provided in 1st row. i.e key and store in list
    for i in sheet[1]:
        data_key.append(i.value)
    data_value = list()
    #  get value in row number provided in row_number and store in list
    for i in sheet[row_number]:
        data_value.append(i.value)

    #  zip both the lists and wrap in dictionary to get key value pair dictonary.
    test_value = dict(zip(data_key, data_value))
    return test_value


def write_data_by_key(file, sheetname, key, row_num, data):
    try:
        wb = op.load_workbook(file)
        sheet = wb[sheetname]
        y = 0
        for i in sheet[1]:
            # cell count to fetch new cell for saving data
            y += 1
            if i.value == key:
                #  Get new cell ce based on passed row number and calculated cell number
                ce = sheet.cell(row=row_num, column=y)
                print(ce)
                ce.value = data
                print(ce.value)
                print('Data is successfully edited')
        wb.save(file)
        print('data is successfully saved')
    except FileNotFoundError:
        print("file not found ")
    except PermissionError:
        print("Permission denied : File cannot be saved")


def write_data_by_row_col(file, sheetname, row_number, col_number, data):
    wb = op.load_workbook(file)
    sheet = wb[sheetname]
    sheet.cell(row=row_number, column=col_number).value = data
    wb.save(file)
    print('Data is successfully saved')