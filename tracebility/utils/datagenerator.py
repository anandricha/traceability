import openpyxl as op
import itertools


def data_generator_return__nested_list(file : str, sheetname : str):
    wb = op.load_workbook(file)
    sheet = wb[sheetname]
    ce = list(sheet.rows)
    print(ce)
    excel_data = list()
    for list_data in ce:

        # list data is the tuple of each row with cell location values
        li1 = list()
        # below for loop is to convert the cell location in tuple to a list of values
        for data in list_data:
            li1.insert(1, data.value)
        # Below code inserts list of values of one row to the main list
        excel_data.insert(1, li1)
        # returing only data so removing first row of keys
    return ()

def data_generator_return_tuple(file : str, sheetname : str):
    wb = op.load_workbook(file)
    sheet = wb[sheetname]
    ce = list(sheet.rows)
    excel_data = list()
    for list_data in ce:
        # list data is the tuple of each row with cell location values
        li1 = list()
        # below for loop is to convert the cell location in tuple to a list of values
        for data in list_data:
            li1.insert(1, data.value)
        # Below code inserts list of values of one row to the main list
        excel_data.append(li1)
        # returing only data so removing first row of keys
    return list(excel_data[1:])


def get_row_data(file,sheetname):
    wb = op.load_workbook(file)
    sheet = wb[sheetname]
    x=0
    for row in sheet.iter_rows(values_only=True):
        if x == 0:
            x=+1
            continue
        yield row